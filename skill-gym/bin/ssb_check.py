#!/usr/bin/env python3
"""SpreadsheetBench gate checker.

Faithful port of the value-comparison path of
RUCKBReasoning/SpreadsheetBench evaluation/evaluation.py (MIT-style academic
release; credited in README). Checks N_output.xlsx against N_answer.xlsx at
answer_position for the three test cases. Hard metric: all 3 must match.

Usage: ssb_check.py <task_dir> <workspace_dir> <answer_position>
Prints JSON: {"passed": bool, "cases": [...], "detail": str}
"""
import datetime
import json
import sys

import openpyxl


def datetime_to_float(dt):
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def transform_value(v):
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        v = round(datetime_to_float(v), 0)
    elif isinstance(v, str):
        try:
            v = round(float(v), 2)
        except ValueError:
            pass
    return v


def compare_cell_value(v1, v2):
    v1 = transform_value(v1)
    v2 = transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) != type(v2):
        return False
    return v1 == v2


def col_num2name(n):
    name = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def col_name2num(name):
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num


def parse_cell_range(range_str):
    start_cell, end_cell = range_str.split(":")
    start_col, start_row, end_col, end_row = "", "", "", ""
    for ch in start_cell:
        (start_row, start_col) = (
            (start_row + ch, start_col) if ch.isdigit() else (start_row, start_col + ch)
        )
    for ch in end_cell:
        (end_row, end_col) = (
            (end_row + ch, end_col) if ch.isdigit() else (end_row, end_col + ch)
        )
    return (col_name2num(start_col), int(start_row)), (col_name2num(end_col), int(end_row))


def generate_cell_names(range_str):
    if ":" not in range_str:
        return [range_str]
    (sc, sr), (ec, er) = parse_cell_range(range_str)
    cols = [col_num2name(i) for i in range(sc, ec + 1)]
    return [f"{c}{r}" for c in cols for r in range(sr, er + 1)]


def split_position(position):
    """'Bank Balance'!AL4:AL21 -> ('Bank Balance', 'AL4:AL21'); M2:M5 -> (None, 'M2:M5')"""
    if "!" in position:
        sheet, rng = position.rsplit("!", 1)
        return sheet.strip("'\""), rng
    return None, position


def check_case(answer_path, output_path, position):
    sheet, rng = split_position(position)
    try:
        wb_gt = openpyxl.load_workbook(answer_path, data_only=True)
    except Exception as e:
        return False, f"cannot open answer: {e}"
    try:
        wb_proc = openpyxl.load_workbook(output_path, data_only=True)
    except Exception as e:
        return False, f"cannot open output: {e}"

    sheet_gt = sheet if sheet and sheet in wb_gt.sheetnames else wb_gt.active.title
    if sheet and sheet in wb_proc.sheetnames:
        sheet_proc = sheet
    elif sheet_gt in wb_proc.sheetnames:
        sheet_proc = sheet_gt
    else:
        return False, f"worksheet {sheet_gt!r} not found in output"

    ws_gt, ws_proc = wb_gt[sheet_gt], wb_proc[sheet_proc]
    for cell in generate_cell_names(rng):
        if not compare_cell_value(ws_gt[cell].value, ws_proc[cell].value):
            return False, (
                f"mismatch at {cell}: expected {ws_gt[cell].value!r}, "
                f"got {ws_proc[cell].value!r}"
            )
    return True, "ok"


def main():
    task_dir, workspace, position = sys.argv[1], sys.argv[2], sys.argv[3]
    import glob
    import os

    cases = []
    for i in (1, 2, 3):
        ans = glob.glob(os.path.join(task_dir, f"{i}_*_answer.xlsx"))
        out = os.path.join(workspace, f"{i}_output.xlsx")
        if not ans:
            cases.append({"case": i, "passed": False, "detail": "answer file missing"})
            continue
        if not os.path.exists(out):
            cases.append({"case": i, "passed": False, "detail": f"{i}_output.xlsx not produced"})
            continue
        ok, detail = check_case(ans[0], out, position)
        cases.append({"case": i, "passed": ok, "detail": detail})

    passed = all(c["passed"] for c in cases)
    print(json.dumps({"passed": passed, "cases": cases,
                      "detail": "hard-pass (3/3)" if passed else "failed"}))
    sys.exit(0)


if __name__ == "__main__":
    main()
